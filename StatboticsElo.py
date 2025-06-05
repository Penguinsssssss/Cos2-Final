import statbotics
sb = statbotics.Statbotics()
import openpyxl
from openpyxl import load_workbook
import sys
from pathlib import Path
import time

def elo(player, enemy, wl, k=32):
    playerExpected = 1 / (1 + 10 ** ((enemy - player) / 400))
    playerActual = 1 if wl else 0
    player += k * (playerActual - playerExpected)
    return round(player, 1)

def main():
    start = time.time()
    teamsElo = {}
    teamsHistory = {}
    history = {}
    id_no = 1
    time_splits = []
    for year in range(2007, 2026):
        try:
            eventsPre = sb.get_events(year=year, limit=999)
        except Exception as e:
            print(f"Error fetching events for {year}: {e}")
            continue
        events = []
        for event in eventsPre:
            events.append([event["key"], event["week"]])
        events.sort(key=lambda x: x[1])
        
        print(events)
        
        ys = time.time()
        for event in events:
            print(event, round(time.time() - start, 1), round(time.time() - ys, 1))
            try:
                matches = sb.get_matches(event=event[0])
            except Exception as e:
                print(f"Error at event {event[0]}: {e}")
                continue
            
            for match in matches:
                
                teamRed = match["alliances"]["red"]["team_keys"]
                teamBlue = match["alliances"]["blue"]["team_keys"]
                winner = match["result"]["winner"]
                
                if winner == "tie": #ties
                    
                    for robot in teamRed + teamBlue:
                        
                        if robot not in teamsElo: #add rookie teams and update match count
                            teamsElo[robot] = {"name": robot, "elo": 1500, "matches": 0, "wins": 0, "losses": 0, "draws": 0}
                            teamsHistory[robot] = []
                        
                        teamsElo[robot]["matches"] += 1
                        teamsElo[robot]["draws"] += 1
                        
                        teamsHistory[robot].append([match["key"], teamsElo[robot]["elo"], "draw"]) #add match to each robots personal history
                    
                    history[match["key"]] = {
                        "victor": "draw",
                        "red": teamRed,
                        "blue": teamBlue,
                        "id_no": id_no
                        }
                    id_no += 1
                
                elif winner == None: #unplayed/bugged matches
                    pass
                
                elif winner == "red" or winner == "blue": #ensure match is valid
                    
                    history[match["key"]] = { #add match to log
                        "victor": winner,
                        "red": teamRed,
                        "blue": teamBlue,
                        "id_no": id_no
                        }
                    id_no += 1
                    
                    
                    for robot in teamRed + teamBlue: #add rookie teams and update match count
                        if robot not in teamsElo:
                            teamsElo[robot] = {"name": robot, "elo": 1500, "matches": 0, "wins": 0, "losses": 0, "draws": 0}
                            teamsHistory[robot] = []
                        teamsElo[robot]["matches"] += 1
                        
                    
                    rwin = (winner == "red")
                    bwin = (winner == "blue")
                    
                    raverage = sum(teamsElo[k]["elo"] for k in teamRed) / len(teamRed)
                    baverage = sum(teamsElo[k]["elo"] for k in teamBlue) / len(teamBlue)
                    
                    for robot in teamRed: #update all red robots
                        if rwin:
                            teamsElo[robot]["wins"] += 1
                            outcome = "win"
                        else:
                            teamsElo[robot]["losses"] += 1
                            outcome = "loss"
                        
                        teamsElo[robot]["elo"] = elo(teamsElo[robot]["elo"], baverage, rwin)
                        
                        teamsHistory[robot].append([match["key"], teamsElo[robot]["elo"], outcome]) #add match to each robots personal history
                    
                    for robot in teamBlue: #update all blue robots
                        if bwin:
                            teamsElo[robot]["wins"] += 1
                            outcome = "win"
                        else:
                            teamsElo[robot]["losses"] += 1
                            outcome = "loss"
                        
                        teamsElo[robot]["elo"] = elo(teamsElo[robot]["elo"], raverage, bwin)
                        
                        teamsHistory[robot].append([match["key"], teamsElo[robot]["elo"], outcome]) #add match to each robots personal history
        time_splits.append([year, time.time() - ys])
    
    #sort
    teamsElo = dict(sorted(teamsElo.items(), key=lambda item: item[1]['elo'], reverse=True))
    teamsHistory = dict(sorted(teamsHistory.items()))
    history = dict(sorted(history.items(), key=lambda item: item[1]['id_no']))
    
    #getdir
    if hasattr(sys, "_MEIPASS"):
        current_dir = Path(sys._MEIPASS)
    else:
        current_dir = Path(__file__).resolve().parent
    
    #create/overwrite file
    filePath = current_dir / "statboticsElo.xlsx"
    if filePath.exists(): filePath.unlink()
    
    #create workbook
    wb = openpyxl.Workbook()
    wb.save(filePath)
    workbook = load_workbook(filePath)
    
    #add elo sheet
    sheet = workbook.create_sheet(title="Elo Table")
    sheet.append(list(teamsElo[list(teamsElo.keys())[0]].keys()))
    for k,v in teamsElo.items():
        row = list(teamsElo[k].values())
        sheet.append(row)
    
    #add history sheet
    sheet = workbook.create_sheet(title="Matches (unordered)")
    sheet.append(["match", "victor", "red1", "red2", "red3", "blue1", "blue2", "blue3"])
    for key, match in history.items():
        row = [key, match["victor"]] + match["red"] + match["blue"]
        sheet.append(row)
    
    #add teamsHistory sheet (match, nelo, outcome)
    sheet = workbook.create_sheet(title="Teams History")
    sheet.append(["team", "match", "elo", "outcome"])

    for team, matches in teamsHistory.items():
        for match in matches:
            sheet.append([team, match[0], match[1], match[2]])
    
    #remove default sheet
    workbook.remove(workbook['Sheet'])
    #save spreadsheet
    workbook.save(filePath)
    
    #happi
    print(f"Elo saved to {filePath} with a final time of {round(time.time() - start, 1)}")
    print("Year splits:")
    for i in time_splits:
        print(i)
    
main()