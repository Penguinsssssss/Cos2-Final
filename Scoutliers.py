#scoutliers11

print("Welcome to Scoutliers")

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import xlsxwriter
import pandas as pd
import os
import sys
import tkinter as tk
from tkinter import filedialog, simpledialog
from pathlib import Path

###UNIVERSAL FUNCTIONS
def zerofyRows(data):
    return [(key, values) for key, values in data if any(x not in (0, 0.0, None) for x in values)]
def transpose_2d_list(raw):
    tr = [[] for _ in range(len(raw[0]))]
    for i in range(len(raw)):
        for j in range(len(raw[0])):
            tr[j].append(raw[i][j])  # Fix: Append to tr[j] instead of tr[i]
    return tr
def rgb_to_hex(r, g, b):
    print(r, g, b)
    # Ensure the values are within the 1-255 range
    if not all(0 <= value <= 255 for value in (r, g, b)):
        raise ValueError("Each RGB value must be between 0 and 255")
    
    # Convert the RGB values to hex and format as a string
    hex_code = f"{r:02X}{g:02X}{b:02X}"
    return hex_code

### INITIAL SPREADSHEET CREATION
#logic
def spreadsheet_to_3d_list(file_path, sheet):
    workbook = openpyxl.load_workbook(file_path)

    sheet_data = []
    worksheet = workbook[sheet]
    
    for columns in worksheet.iter_cols(values_only=True): #valuesonly does not work ;(
        sheet_data.append(list(columns))

    return sheet_data
def getAvgs(data, ind, name):
    stats = {}
    #print(index)
    #print(data[1])
    whereismyrobot = ind.index(name)
    for i in data:
        if i[whereismyrobot] not in stats:
            stats[i[whereismyrobot]] = i
        else:
            for j in range(len(i)):
                if i[j] is None:
                    continue
                if stats[i[whereismyrobot]][j] is None:
                    stats[i[whereismyrobot]][j] = i[j]
                elif isinstance(i[j], (int, float)) and isinstance(stats[i[whereismyrobot]][j], (int, float)):
                    stats[i[whereismyrobot]][j] = (stats[i[whereismyrobot]][j] + i[j]) / 2
    return stats
#viewers
def viewStatsList(stats):
    for key, values in stats.items():
        print(f"{key}: {values}")
def viewStat(index, stats):
    print()
    #viewStatsList(stats)
    for i in range(len(index)):
        print(i, index[i])
    response = int(input("\n" + "Enter what stat would you like to see: "))
    subStats = []
    for key, values in stats.items():
        subStats.append([key, values[response]])
    subStats = sorted(subStats, key=lambda x: x[1] if x[1] is not None else float('inf'))
    subStats.reverse()
    for i in subStats:
        print(i[0], i[1])
def viewRobot(index, stats):
    #print(stats)
    print()
    robots = []
    for key, values in stats.items():
        if key != None:
            robots.append(int(key))
    robots = sorted([x for x in robots if x is not None])
    for i in robots:
        print(i)
    robot = input("Enter the team number of the robot you want to view: ")
    robot = int(robot)
    if robot in robots:
        print(index)
        for i in range(len(index)):
            print(index[i], stats[robot][i])
#output
def saveToSpreadsheet(index, stats, name, chosenName):
    workbook = xlsxwriter.Workbook(name)
    worksheet = workbook.add_worksheet(chosenName)
    
    sl = [[key, value] for key, value in stats.items()]
    sl = zerofyRows(sl)
    sl.sort(key=lambda x: x[0])
    
    for i in range(len(index)):
        worksheet.write(0, i + 1, index[i])
    for i in range(len(sl)):
        worksheet.write(i + 1, 0, sl[i][0])
        for j in range(len(sl[i][1])):
            worksheet.write(i + 1, j + 1, sl[i][1][j])
    workbook.close()
#main
def createInitial(filepath, sheet, location, name, chosenName, loca):
    data = spreadsheet_to_3d_list(filepath, sheet)
    data = prune(data)
    data = zerofy(data)
    data = transpose_2d_list(data)
    index = data.pop(0)
    stats = getAvgs(data, index, location)
    saveToSpreadsheet(index, stats, name, chosenName)
    print("File of averages saved to " + loca + name)
#debug
def prune(raw):
    rawr = raw
    filtered_data = []
    
    for i in range(len(rawr)):
        good = True
        for j in range(len(rawr[i])):
            if isinstance(rawr[i][j], str) == True and j != 0:
                good = False
        if good:
            filtered_data.append(rawr[i])
    
    #print(filtered_data)
    return filtered_data
def makeIndex(raw):
    ind = [row[0] for row in raw]  # Extract first column
    rawr = [row[1:] for row in raw]  # Remove first element from each row
    print(rawr)
    print(ind)
    return rawr, ind
def zerofy(raw):
    rawr = raw
    for i in range(len(rawr)):
        for j in range(len(rawr[i])):
            if rawr[i][j] is None:
                rawr[i][j] = 0
    return rawr

### TEAM VISUALIZER
def createAlli(filepath, sheetname, robot, name, avg):
    data = spreadsheet_to_3d_list(filepath, sheetname)
    data = transpose_2d_list(data)
    index = data.pop(0)
    stats = getAlliStats(data, robot, avg)
    tsave_2dl_to_sheet(filepath, name, stats, index)
    print("File of possible alliences saved to", filepath)
def tsave_2dl_to_sheet(filepath, sheet_name, stats, index):
    workbook = load_workbook(filepath)
    
    if sheet_name not in workbook.sheetnames:
        sheet = workbook.create_sheet(title=sheet_name)
    else:
        sheet = workbook[sheet_name]
    
    indexx = [""] * 2 + index
    sheet.append(indexx)
    
    for i in range(len(stats)):
        fr = []
        
        for robot_num in stats[i][0]:
            fr.append(robot_num)
        
        for stat in stats[i][1]:
            fr.append(stat)
        
        sheet.append(fr)
    
    workbook.save(filepath)
def getAlliStats(data, ourRobot=5687, avr=False):
    new_stats = []
    ourRobot = int(ourRobot)
        
    our_robot_stats = None
    for row in data:
        if row[0] == ourRobot:
            our_robot_stats = row[1:]
            break
    
    if avr:
        a = 3
    else:
        a = 1
    
    for i in range(len(data) - 1):
        for j in range(i + 1, len(data)):
            if ourRobot not in (data[i][0], data[j][0]):
                combined_stats = [
                    (data[i][k] + data[j][k] + our_robot_stats[k - 1]) / a
                    for k in range(1, len(data[i]))
                ]
                new_stats.append([[ourRobot, data[i][0], data[j][0]], combined_stats])

    return new_stats

### PERCENTILE RANKS
def createNorm(filepath, name, chosenName):
    data = spreadsheet_to_3d_list(filepath, chosenName)
    data = transpose_2d_list(data) #right side
    index = data.pop(0)
    data = transpose_2d_list(data) #turned
    stats = getNorms(data)
    stats = transpose_2d_list(stats) #right side
    rsave_2dl_to_sheet(filepath, name, stats, index)
    print("File of normalized averages saved to", filepath)
def getNorms(data):
    if not data:  # Handle empty input
        return []

    stats = [data[0]]  # Keep the first row unchanged
    for sublist in data[1:]:
        if len(sublist) == 0:  # Handle empty sublists
            stats.append([])
            continue

        t = max(sublist)
        b = min(sublist)

        # Normalize only if t != b, otherwise return all zeros
        if t == b:
            normalized = [0] * len(sublist)
        else:
            normalized = [((x - b) / (t - b) * 100) for x in sublist]

        stats.append(normalized)

    return stats
def rsave_2dl_to_sheet(filepath, sheet_name, stats, index):
    workbook = load_workbook(filepath)
    
    if sheet_name not in workbook.sheetnames:
        sheet = workbook.create_sheet(title=sheet_name)
    else:
        sheet = workbook[sheet_name]
    
    sheet.append(index)
    
    for i in range(len(stats)):
        sheet.append(stats[i])
        for j in range(1, len(stats[i])):
            if stats[i][j] <= 50:
                color = rgb_to_hex(255, int(255 * (stats[i][j] / 50)), 0)
            else:
                color = rgb_to_hex(int(255 * (1 - (stats[i][j] - 50) / 50)), 255, 0)
            sheet.cell(row=i+2, column=j+1).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    
    workbook.save(filepath)

### PICKS GUIDE
def createGuide(filepath, name, chosenName):
    data = spreadsheet_to_3d_list(filepath, chosenName)
    data = transpose_2d_list(data)
    index = data.pop(0)
    index = index[3:]
    data = transpose_2d_list(data)
    bstats = getGuide(data, True)
    bstats = transpose_2d_list(bstats)
    wstats = getGuide(data, False)
    wstats = transpose_2d_list(wstats)
    gsave_2dl_to_sheet(filepath, name, bstats, wstats, index)
    print("File of picks saved to", filepath)
def getGuide(data, borw):
    stats = []
    headers = data[:3]  # Assuming first 3 rows are headers
    
    for i in range(3, len(data)):  # Starting from row 4 onward
        row = data[i]
        if borw:
            highest_stat = max(row)
        else:
            highest_stat = min(row)
        highest_header_index = row.index(highest_stat)  # Get the index of the highest stat
        header = [headers[j][highest_header_index] for j in range(3)]  # Find the corresponding header(s)
        
        stats.append(header)  # Append the stat and headers tuple
        
    return stats
def combineStats(b, w):
    stats = []
    stats.append(["Best Possible Aliances"])
    for i in b:
        stats.append(b)
    stats.append(["Worst Possible Aliances"])
    for i in w:
        stats.append(w)
    for i in range(len(stats)):
        for j in range(len(stats[i])):
            stats[i][j] = str(stats[i][j])
    print(stats)
    return stats
def gsave_2dl_to_sheet(filepath, sheet_name, bstats, wstats, index):
    workbook = load_workbook(filepath)
    
    if sheet_name not in workbook.sheetnames:
        sheet = workbook.create_sheet(title=sheet_name)
    else:
        sheet = workbook[sheet_name]
    
    sheet.append(index)
    sheet.append(["Best Possible Alliances"])
    for i in range(len(bstats)):
        sheet.append(bstats[i])
    sheet.append(["Worst Possible Alliances"])
    for i in range(len(wstats)):
        sheet.append(wstats[i])
    
    workbook.save(filepath)

### FILE SELECT
def fileSelect():
    # Open file dialog to select Excel file
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(title="Select an Excel File", filetypes=[("Excel files", "*.xlsx")])
    if not file_path:
        print("No file selected. Exiting.")
        exit()
    xls = pd.ExcelFile(file_path)

    # Prompt for sheet selection
    sheet_name = simpledialog.askstring("Select the sheet containing raw data: ", f"Available sheets:\n{xls.sheet_names}\nEnter sheet name:")
    if sheet_name not in xls.sheet_names:
        print("Invalid sheet name. Exiting.")
        exit()
    df = pd.read_excel(file_path, sheet_name=sheet_name)

    # Prompt for column selection
    column_name = simpledialog.askstring("Select the column containing team numbers", f"Available columns:\n{df.columns.tolist()}\nEnter column name:")
    if column_name not in df.columns:
        print("Invalid column name. Exiting.")
        exit()
    
    return file_path, sheet_name, column_name

### Program Input
inputFilepath = "/Users/BenjaminSullivan/Downloads/Pine Tree Scouting - Week 3.xlsx"
inputSheetpath = "Raw data" #Sheet containing raw data
inputRobotpath = "Team Number" #Column of sheet containing team numbers
inputFilepath, inputSheetpath, inputRobotpath = fileSelect()

### Program Output
"""if os.name == "nt":  # Windows
    #outputLocation = Path.home() / f"Downloads"  # This is a Path object
    outputLocation = Path.home() / "Downloads"
    print(outputLocation)
else:  # macOS/Linux
    outputLocation = os.path.join(os.path.expanduser("~"), "Downloads")"""

if getattr(sys, 'frozen', False):  # Check if running as an .exe
    outputLocation = Path(sys.executable).parent
else:
    outputLocation = Path.cwd()
outputLocation = str(outputLocation)


outputName = "\\" + simpledialog.askstring("Name Selection", "Name the output file: ") + ".xlsx"

fullPath = outputLocation + outputName
print(f"Current working directory: {os.getcwd()}")

print(f"Saving to: {fullPath}")
#outputName = "Colored Output 1.xlsx"

### Customization
rawStatsName = "Indiv Stats"
rawAlliStatsName = "Alliance Stats"
rankStatsName = "Normalized Indiv Stats"
rankAlliStatsName = "Normalized Alliance Stats (avg)"
guideName = "Picks Guide"

### Team Number
teamNumber = 5687

createInitial(inputFilepath, inputSheetpath, inputRobotpath, fullPath, rawStatsName, outputLocation)
#createAlli(fullPath, rawStatsName, teamNumber, rawAlliStatsName, False)
createNorm(fullPath, rankStatsName, rawStatsName)
#createAlli(fullPath, rankStatsName, teamNumber, rankAlliStatsName, True)
#createGuide(fullPath, guideName, rankAlliStatsName)

#Pine Tree Scouting - Week 3.xlsx
#Week 1 Scouting Sheet 2_29-3_2 .xlsx